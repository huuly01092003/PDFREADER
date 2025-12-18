import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers

# File Excel riêng cho format 2
EXCEL_FILE_FORMAT2 = os.path.join(os.path.dirname(__file__), "output_format2.xlsx")

def init_excel_format2():
    '''Khởi tạo file Excel format 2 với number formatting'''
    if os.path.exists(EXCEL_FILE_FORMAT2):
        try:
            wb = load_workbook(EXCEL_FILE_FORMAT2)
            ws = wb.active
            
            if ws.max_row >= 1:
                first_row = [cell.value for cell in ws[1]]
                expected_headers = [
                    "ThoiGianThucThi", "FileName", "OrderNo", "OrderDate",
                    "SupplierCode", "ComContract", "OrderedBy", "DeliveredTo",
                    "ForStore", "Article", "ArticleDesc", "OUType", "LV",
                    "SKU_OU", "OUQty", "FreeQty", "NetPurchasePrice",
                    "Unit", "TotalNetPurchasePrice"
                ]
                
                if first_row == expected_headers:
                    wb.close()
                    return True
            
            wb.close()
        except Exception as e:
            print(f"File Excel lỗi: {e}, tạo lại...")
            try:
                os.remove(EXCEL_FILE_FORMAT2)
            except:
                pass
    
    # Tạo file mới
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Orders"
    
    # Headers
    headers = [
        "ThoiGianThucThi", "FileName", "OrderNo", "OrderDate",
        "SupplierCode", "ComContract", "OrderedBy", "DeliveredTo",
        "ForStore", "Article", "ArticleDesc", "OUType", "LV",
        "SKU_OU", "OUQty", "FreeQty", "NetPurchasePrice",
        "Unit", "TotalNetPurchasePrice"
    ]
    
    ws.append(headers)
    
    # Style cho header
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F9D58", end_color="0F9D58", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Set column widths
    column_widths = {
        'A': 18,  # ThoiGianThucThi
        'B': 30,  # FileName
        'C': 15,  # OrderNo
        'D': 12,  # OrderDate
        'E': 12,  # SupplierCode
        'F': 12,  # ComContract
        'G': 40,  # OrderedBy
        'H': 40,  # DeliveredTo
        'I': 40,  # ForStore
        'J': 15,  # Article
        'K': 40,  # ArticleDesc
        'L': 10,  # OUType
        'M': 8,   # LV
        'N': 10,  # SKU_OU
        'O': 10,  # OUQty
        'P': 10,  # FreeQty
        'Q': 15,  # NetPurchasePrice
        'R': 8,   # Unit
        'S': 18   # TotalNetPurchasePrice
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = 'A2'
    
    wb.save(EXCEL_FILE_FORMAT2)
    print(f"✅ Đã tạo file Excel Format 2 với {len(headers)} cột")
    return True

def get_existing_records_format2():
    '''Lấy danh sách các record đã tồn tại (Format 2)'''
    if not os.path.exists(EXCEL_FILE_FORMAT2):
        return set()
    
    try:
        wb = load_workbook(EXCEL_FILE_FORMAT2)
        ws = wb.active
        
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] and row[2] and row[9]:  # FileName, OrderNo, Article
                key = f"{row[1]}|{row[2]}|{row[9]}"
                existing.add(key)
        
        wb.close()
        return existing
        
    except Exception as e:
        print(f"Lỗi đọc existing records: {e}")
        return set()

def append_excel_format2(rows):
    '''Thêm dữ liệu vào Excel Format 2 với number formatting'''
    try:
        if not os.path.exists(EXCEL_FILE_FORMAT2):
            init_excel_format2()
        
        existing_records = get_existing_records_format2()
        
        wb = load_workbook(EXCEL_FILE_FORMAT2)
        ws = wb.active
        
        added_count = 0
        skipped_count = 0
        
        for r in rows:
            # Check duplicate: filename|orderno|article
            if len(r) >= 10:
                key = f"{r[1]}|{r[2]}|{r[9]}"
                
                if key in existing_records:
                    skipped_count += 1
                    continue
                
                existing_records.add(key)
            
            # Append row
            ws.append(r)
            added_count += 1
            
            # Apply number formatting to the just-added row
            current_row = ws.max_row
            
            # Column M: LV (index 12) - can have decimals like 1.5
            cell_m = ws[f'M{current_row}']
            if isinstance(r[12], (int, float)):
                cell_m.value = r[12]
                if r[12] % 1 == 0:  # Integer
                    cell_m.number_format = '0'
                else:  # Has decimals
                    cell_m.number_format = '0.0'
            
            # Column N: SKU_OU (index 13) - can have decimals
            cell_n = ws[f'N{current_row}']
            if isinstance(r[13], (int, float)):
                cell_n.value = r[13]
                if r[13] % 1 == 0:
                    cell_n.number_format = '0'
                else:
                    cell_n.number_format = '0.0'
            
            # Column O: OUQty (index 14) - can have decimals
            cell_o = ws[f'O{current_row}']
            if isinstance(r[14], (int, float)):
                cell_o.value = r[14]
                if r[14] % 1 == 0:
                    cell_o.number_format = '0'
                else:
                    cell_o.number_format = '0.0'
            
            # Column P: FreeQty (index 15) - can have decimals
            cell_p = ws[f'P{current_row}']
            if isinstance(r[15], (int, float)):
                cell_p.value = r[15]
                if r[15] % 1 == 0:
                    cell_p.number_format = '0'
                else:
                    cell_p.number_format = '0.0'
            
            # Column Q: NetPurchasePrice (index 16) - always integer (70.000 → 70000)
            cell_q = ws[f'Q{current_row}']
            if isinstance(r[16], (int, float)):
                cell_q.value = r[16]
                cell_q.number_format = '0'  # Integer format, no decimals
            
            # Column S: TotalNetPurchasePrice (index 18) - always integer
            cell_s = ws[f'S{current_row}']
            if isinstance(r[18], (int, float)):
                cell_s.value = r[18]
                cell_s.number_format = '0'  # Integer format, no decimals
        
        wb.save(EXCEL_FILE_FORMAT2)
        
        if skipped_count > 0:
            print(f"ℹ️ [Format 2] Đã thêm {added_count} dòng, bỏ qua {skipped_count} dòng trùng")
        
        return True
        
    except Exception as e:
        print(f"❌ Lỗi append Excel Format 2: {e}")
        import traceback
        traceback.print_exc()
        return False

def read_excel_data_format2():
    '''Đọc dữ liệu từ Excel Format 2'''
    if not os.path.exists(EXCEL_FILE_FORMAT2):
        print("⚠️ File Excel Format 2 không tồn tại, tạo mới...")
        init_excel_format2()
        return []
    
    try:
        wb = load_workbook(EXCEL_FILE_FORMAT2)
        ws = wb.active
        
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                data.append(row)
        
        wb.close()
        return data
        
    except Exception as e:
        print(f"❌ Lỗi đọc Excel Format 2: {e}")
        try:
            os.remove(EXCEL_FILE_FORMAT2)
        except:
            pass
        init_excel_format2()
        return []

def clear_excel_data_format2():
    '''Xóa tất cả dữ liệu Format 2'''
    try:
        if not os.path.exists(EXCEL_FILE_FORMAT2):
            init_excel_format2()
            return True
        
        wb = load_workbook(EXCEL_FILE_FORMAT2)
        ws = wb.active
        
        ws.delete_rows(2, ws.max_row)
        
        wb.save(EXCEL_FILE_FORMAT2)
        print("🗑️ Đã xóa dữ liệu Excel Format 2 (giữ header)")
        return True
        
    except Exception as e:
        print(f"❌ Lỗi xóa Excel Format 2: {e}")
        return False