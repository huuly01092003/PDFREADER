import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import EXCEL_FILE

def init_excel():
    '''Khởi tạo file Excel với header đầy đủ'''
    if os.path.exists(EXCEL_FILE):
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            
            if ws.max_row >= 1:
                first_row = [cell.value for cell in ws[1]]
                expected_headers = [
                    "ThoiGianThucThi", "FileName", "PONumber", "SKUNumber",
                    "Description", "Vendor", "PartNo",
                    "BuyCost", "NetBuyCost", 
                    "QtyOrdCS", "QtyOrdPcs", "QtyRecPcs",
                    "ExtendedCost"
                ]
                
                if first_row == expected_headers:
                    wb.close()
                    return True
            
            wb.close()
        except Exception as e:
            print(f"File Excel lỗi: {e}, tạo lại...")
            try:
                os.remove(EXCEL_FILE)
            except:
                pass
    
    # Tạo file mới với header đầy đủ
    wb = Workbook()
    ws = wb.active
    ws.title = "DATA"
    
    # Headers đầy đủ
    headers = [
        "ThoiGianThucThi", "FileName", "PONumber", "SKUNumber",
        "Description", "VendorPartNo", "SellUM", "BuyUM",
        "BuyCost", "NetBuyCost", 
        "QtyOrdCS", "QtyOrdPcs", "QtyRecPcs",
        "ExtendedCost"
    ]
    
    ws.append(headers)
    
    # Style cho header
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Set column widths
    column_widths = {
        'A': 18,  # ThoiGianThucThi
        'B': 25,  # FileName
        'C': 15,  # PONumber
        'D': 15,  # SKUNumber
        'E': 40,  # Description
        'F': 10,  # Vendor
        'G': 10,  # PartNo
        'H': 12,  # BuyCost
        'I': 12,  # NetBuyCost
        'J': 12,  # QtyOrdCS
        'K': 12,  # QtyOrdPcs
        'L': 12,  # QtyRecPcs
        'M': 15   # ExtendedCost
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = 'A2'
    
    wb.save(EXCEL_FILE)
    print(f"✅ Đã tạo file Excel mới với {len(headers)} cột")
    return True

def get_existing_records():
    '''Lấy danh sách các record đã tồn tại'''
    if not os.path.exists(EXCEL_FILE):
        return set()
    
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] and row[2] and row[3]:  # FileName, PONumber, SKUNumber
                key = f"{row[1]}|{row[2]}|{row[3]}"
                existing.add(key)
        
        wb.close()
        return existing
        
    except Exception as e:
        print(f"Lỗi đọc existing records: {e}")
        return set()

def append_excel(rows):
    '''Thêm dữ liệu vào Excel'''
    try:
        if not os.path.exists(EXCEL_FILE):
            init_excel()
        
        existing_records = get_existing_records()
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        added_count = 0
        skipped_count = 0
        
        for r in rows:
            # Check duplicate: filename|po|sku
            if len(r) >= 4:
                key = f"{r[1]}|{r[2]}|{r[3]}"
                
                if key in existing_records:
                    skipped_count += 1
                    continue
                
                existing_records.add(key)
            
            ws.append(r)
            added_count += 1
        
        wb.save(EXCEL_FILE)
        
        if skipped_count > 0:
            print(f"ℹ️ Đã thêm {added_count} dòng, bỏ qua {skipped_count} dòng trùng")
        
        return True
        
    except Exception as e:
        print(f"❌ Lỗi append Excel: {e}")
        return False

def read_excel_data():
    '''Đọc dữ liệu từ Excel'''
    if not os.path.exists(EXCEL_FILE):
        print("⚠️ File Excel không tồn tại, tạo mới...")
        init_excel()
        return []
    
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                data.append(row)
        
        wb.close()
        return data
        
    except Exception as e:
        print(f"❌ Lỗi đọc Excel: {e}")
        try:
            os.remove(EXCEL_FILE)
        except:
            pass
        init_excel()
        return []

def clear_excel_data():
    '''Xóa tất cả dữ liệu'''
    try:
        if not os.path.exists(EXCEL_FILE):
            init_excel()
            return True
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        ws.delete_rows(2, ws.max_row)
        
        wb.save(EXCEL_FILE)
        print("🗑️ Đã xóa dữ liệu Excel (giữ header)")
        return True
        
    except Exception as e:
        print(f"❌ Lỗi xóa Excel: {e}")
        return False

def get_excel_stats():
    '''Lấy thống kê Excel'''
    if not os.path.exists(EXCEL_FILE):
        return {
            'total_rows': 0,
            'total_files': 0,
            'total_pos': 0,
            'file_exists': False
        }
    
    try:
        data = read_excel_data()
        
        files = set()
        pos = set()
        
        for row in data:
            if len(row) >= 3:
                files.add(row[1])
                pos.add(row[2])
        
        return {
            'total_rows': len(data),
            'total_files': len(files),
            'total_pos': len(pos),
            'file_exists': True
        }
        
    except:
        return {
            'total_rows': 0,
            'total_files': 0,
            'total_pos': 0,
            'file_exists': False
        }